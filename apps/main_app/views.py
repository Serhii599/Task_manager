from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponseRedirect, HttpResponse
from django.shortcuts import render, redirect
from django.urls import reverse_lazy
from django.views.generic import View, ListView, TemplateView, CreateView, UpdateView, DeleteView, DetailView
from django.shortcuts import get_object_or_404
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.utils import timezone

from .forms import TaskCreationForm, ProjectCreationForm, CommentForm
from .models import *

User = get_user_model()


class MainView(View):
    def get(self, request):
        if request.user.is_authenticated:
            return redirect('main_app:my_tasks')
        else:
            return redirect('authentication:login')


class MyTasksListView(LoginRequiredMixin, ListView):
    model = Task
    context_object_name = 'tasks'
    template_name = 'main_app/my_task_list.html'
    
    paginate_by = 30

    def get_queryset(self):
        qs = Task.objects.filter(assignee=self.request.user)

        req = self.request.GET

        # FILTERS ------------------------------------------------------
        status = req.get("status")
        priority = req.get("priority")
        project_id = req.get("project")
        date_from = req.get("date_from")
        date_to = req.get("date_to")

        if status:
            qs = qs.filter(status=status)

        if priority:
            qs = qs.filter(priority=priority)

        if project_id:
            qs = qs.filter(projects__id=project_id)

        if date_from:
            qs = qs.filter(due_date__gte=date_from)

        if date_to:
            qs = qs.filter(due_date__lte=date_to)

        # SORTING ------------------------------------------------------
        sort_field = req.get("sort")
        sort_order = req.get("order", "asc")

        sortable_fields = [
            "task_name", "assignee", "task_description",
            "status", "priority", "projects.project_name", "due_date"
        ]

        if sort_field in sortable_fields:
            if sort_order == "desc":
                qs = qs.order_by(f"-{sort_field}")
            else:
                qs = qs.order_by(sort_field)

        return qs

    # =================================================================
    # CONTEXT: передаємо sort_columns = [{field, label, url}]
    # =================================================================
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        request = self.request
        params = request.GET.copy()

        sort_definitions = {
            "task_name": "Name",
            "assignee": "Assignee",
            "task_description": "Description",
            "status": "Status",
            "priority": "Priority",
            "project_name": "Project",
            "due_date": "Due date",
        }

        current_sort = params.get("sort")
        current_order = params.get("order", "asc")

        sort_columns = []

        for field, label in sort_definitions.items():
            new_params = params.copy()

            # Toggle logic
            if current_sort == field and current_order == "asc":
                new_params["order"] = "desc"
            else:
                new_params["order"] = "asc"

            new_params["sort"] = field

            sort_columns.append({
                "field": field,
                "label": label,
                "url": "?" + new_params.urlencode()
            })

        context["sort_columns"] = sort_columns

        context["projects"] = Project.objects.all()
        context["assignees"] = User.objects.filter(task__isnull=False).distinct()

        return context



class OneTaskDetailView(LoginRequiredMixin, DetailView):
    model = Task
    context_object_name = 'task'
    template_name = 'main_app/one_task.html'
    pk_url_kwarg = 'task_id'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        task = self.object  # поточна задача

        # знаходимо інші задачі того самого виконавця
        context["other_tasks"] = (
            Task.objects
            .filter(assignee=task.assignee)
            .exclude(id=task.id)
            .order_by("due_date")
        )
        
        # Check if user is collaborator or assignee or creator
        user = self.request.user
        is_collaborator = (
            user == task.assignee or 
            user == task.creator or 
            user in task.collaborators.all()
        )
        context["is_collaborator"] = is_collaborator
        
        # Get comments
        context["comments"] = task.comments.all()
        
        # Comment form
        if is_collaborator:
            context["comment_form"] = CommentForm()

        return context

class ProjectsListView(LoginRequiredMixin, ListView):
    model = Project
    context_object_name = 'projects'
    template_name = 'main_app/projects_list.html'

    def get_queryset(self):
        qs = Project.objects.all()

        req = self.request.GET

        # FILTERS ------------------------------------------------------
        status = req.get("status")
        priority = req.get("priority")
        creator_id = req.get("creator")

        if status:
            qs = qs.filter(status=status)

        if priority:
            qs = qs.filter(priority=priority)

        if creator_id:
            qs = qs.filter(creator_id=creator_id)

        # SORTING ------------------------------------------------------
        sort_field = req.get("sort")
        sort_order = req.get("order", "asc")

        sortable_fields = [
            "project_name", "status", "priority", "task_count", "creator__first_name", "created_at"
        ]

        if sort_field in sortable_fields:
            if sort_order == "desc":
                qs = qs.order_by(f"-{sort_field}")
            else:
                qs = qs.order_by(sort_field)

        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        request = self.request
        params = request.GET.copy()

        sort_definitions = {
            "project_name": "Name",
            "status": "Status",
            "priority": "Priority",
            "task_count": "Tasks number",
            "creator__first_name": "Creator",
            "created_at": "Created at",
        }

        current_sort = params.get("sort")
        current_order = params.get("order", "asc")

        sort_columns = []

        for field, label in sort_definitions.items():
            new_params = params.copy()

            # Toggle logic
            if current_sort == field and current_order == "asc":
                new_params["order"] = "desc"
            else:
                new_params["order"] = "asc"

            new_params["sort"] = field

            sort_columns.append({
                "field": field,
                "label": label,
                "url": "?" + new_params.urlencode()
            })

        context["sort_columns"] = sort_columns
        context["creators"] = User.objects.filter(created_project__isnull=False).distinct()

        return context


class OneProjectListView(LoginRequiredMixin, DetailView):
    model = Project
    context_object_name = 'project'
    template_name = 'main_app/one_project.html'
    pk_url_kwarg = 'project_id'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        project = self.object
        tasks = Task.objects.filter(projects=project)

        # --- FILTERING ---
        status = self.request.GET.get("status")
        priority = self.request.GET.get("priority")
        assignee_id = self.request.GET.get("assignee")

        if status:
            tasks = tasks.filter(status=status)

        if priority:
            tasks = tasks.filter(priority=priority)

        if assignee_id:
            tasks = tasks.filter(assignee_id=assignee_id)

        # --- SORTING ---
        sort_field = self.request.GET.get("sort")
        sort_order = self.request.GET.get("order", "asc")

        sortable_fields = [
            "task_name", "due_date", "status", "priority", "assignee__first_name"
        ]

        if sort_field in sortable_fields:
            if sort_order == "desc":
                tasks = tasks.order_by(f"-{sort_field}")
            else:
                tasks = tasks.order_by(sort_field)

        # Build sort_columns for template
        request = self.request
        params = request.GET.copy()

        sort_definitions = {
            "task_name": "Name",
            "due_date": "Due date",
            "assignee__first_name": "Assignee",
            "status": "Status",
            "priority": "Priority",
        }

        current_sort = params.get("sort")
        current_order = params.get("order", "asc")

        sort_columns = []

        for field, label in sort_definitions.items():
            new_params = params.copy()

            # Toggle logic
            if current_sort == field and current_order == "asc":
                new_params["order"] = "desc"
            else:
                new_params["order"] = "asc"

            new_params["sort"] = field

            sort_columns.append({
                "field": field,
                "label": label,
                "url": "?" + new_params.urlencode()
            })

        context["tasks"] = tasks
        context["sort_columns"] = sort_columns
        context["assignees"] = User.objects.filter(task__isnull=False).distinct()
        return context

class ProjectCreateView(LoginRequiredMixin, CreateView):
    model = Project
    form_class = ProjectCreationForm
    template_name = "main_app/project_create.html"
    success_url = reverse_lazy("main_app:project_create")

    def form_valid(self, form):
        form.instance.creator = self.request.user
        return super().form_valid(form)


class TaskCreateView(LoginRequiredMixin, CreateView):
    model = Task
    form_class = TaskCreationForm
    template_name = "main_app/task_create.html"
    success_url = reverse_lazy("main_app:task_create")

    def form_valid(self, form):
        # приклад, якщо треба підв’язати користувача
        form.instance.creator = self.request.user
        return super().form_valid(form)


class ProjectUpdateView(LoginRequiredMixin, UpdateView):
    model = Project
    form_class = ProjectCreationForm
    template_name = 'main_app/project_edit.html'
    pk_url_kwarg = 'project_id'

    def get_success_url(self):
        return reverse_lazy("main_app:one_project", kwargs={"project_id": self.object.id})


class TaskUpdateView(LoginRequiredMixin, UpdateView):
    model = Task
    form_class = TaskCreationForm
    template_name = 'main_app/task_edit.html'
    pk_url_kwarg = 'task_id'

    def get_success_url(self):
        return reverse_lazy("main_app:one_task", kwargs={"task_id": self.object.id})

class ProjectDeleteView(LoginRequiredMixin, DeleteView):
    model = Project
    pk_url_kwarg = 'project_id'
    template_name = "main_app/project_confirm_delete.html"
    success_url = reverse_lazy("main_app:projects_view")


class TaskDeleteView(LoginRequiredMixin, DeleteView):
    model = Task
    success_url = reverse_lazy('main_app:my_tasks')

class UsersListView(LoginRequiredMixin, ListView):
    model = User
    context_object_name = 'users'
    template_name = 'main_app/users_list.html'

class UserTasksView(LoginRequiredMixin, ListView):
    model = Task
    context_object_name = 'tasks'
    template_name = 'main_app/user_tasks.html'

    def get_queryset(self):
        user_id = self.kwargs['user_id']
        qs = Task.objects.filter(assignee_id=user_id).select_related('assignee')

        req = self.request.GET

        # FILTERS ------------------------------------------------------
        status = req.get("status")
        priority = req.get("priority")

        if status:
            qs = qs.filter(status=status)

        if priority:
            qs = qs.filter(priority=priority)

        # SORTING ------------------------------------------------------
        sort_field = req.get("sort")
        sort_order = req.get("order", "asc")

        sortable_fields = [
            "task_name", "task_description", "status", "priority", "due_date", "created_at"
        ]

        if sort_field in sortable_fields:
            if sort_order == "desc":
                qs = qs.order_by(f"-{sort_field}")
            else:
                qs = qs.order_by(sort_field)

        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user_id = self.kwargs['user_id']

        context['user'] = get_object_or_404(User, id=user_id)

        request = self.request
        params = request.GET.copy()

        sort_definitions = {
            "task_name": "Name",
            "task_description": "Description",
            "status": "Status",
            "priority": "Priority",
            "due_date": "Due date",
            "created_at": "Created at",
        }

        current_sort = params.get("sort")
        current_order = params.get("order", "asc")

        sort_columns = []

        for field, label in sort_definitions.items():
            new_params = params.copy()

            # Toggle logic
            if current_sort == field and current_order == "asc":
                new_params["order"] = "desc"
            else:
                new_params["order"] = "asc"

            new_params["sort"] = field

            sort_columns.append({
                "field": field,
                "label": label,
                "url": "?" + new_params.urlencode()
            })
        context["sort_columns"] = sort_columns
        context["assignees"] = User.objects.filter(task__isnull=False).distinct()

        return context

class TaskMarkDoneView(LoginRequiredMixin, View):
    def post(self, request, task_id):
        task = get_object_or_404(Task, id=task_id)

        # Дозволити змінювати тільки виконавцю або автору
        if task.assignee != request.user and task.creator != request.user:
            return HttpResponse("Forbidden", status=403)

        task.status = "Done"
        task.save()

        return redirect('main_app:one_task', task_id=task.id)
    
class LeaveTaskView(LoginRequiredMixin, View):
    def post(self, request, task_id):
        task = get_object_or_404(Task, id=task_id)
        user = request.user

        # Якщо юзер є асайні — знімаємо його
        if task.assignee == user:
            task.assignee = None

        # Якщо юзер є серед колабораторів — видаляємо
        if user in task.collaborators.all():
            task.collaborators.remove(user)

        task.save()
        return redirect("main_app:my_tasks")
    
class ProjectReportView(LoginRequiredMixin, DetailView):
    model = Project
    template_name = "main_app/project_report.html"
    context_object_name = "project"
    pk_url_kwarg = "project_id"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        project = self.object
        tasks = Task.objects.filter(projects=project)

        # ===========================
        # 1. TOP 3 BY DEADLINE
        # ===========================
        top_by_deadline = tasks.exclude(due_date=None).order_by("due_date")[:3]

        # ===========================
        # 2. TOP 3 BY PRIORITY (custom ordering)
        # ===========================
        priority_order = {"Urgent": 1, "High": 2, "Medium": 3, "Low": 4}

        top_by_priority = sorted(
            tasks,
            key=lambda t: priority_order.get(t.priority, 999)
        )[:3]

        # ===========================
        # 3. OVERDUE TASKS
        # ===========================
        from django.utils import timezone
        now = timezone.now()

        overdue_tasks = tasks.filter(
            due_date__lt=now
        ).exclude(
            status="Done"
        ).order_by("due_date")

        # ===========================
        # 4. TOTAL DONE TASKS
        # ===========================
        total_done = tasks.filter(status="Done").count()

        # ===========================
        # Add to context
        # ===========================
        context["top_by_deadline"] = top_by_deadline
        context["top_by_priority"] = top_by_priority
        context["overdue_tasks"] = overdue_tasks
        context["total_done"] = total_done

        return context

class ProjectReportExcelView(LoginRequiredMixin, View):
    def get(self, request, project_id):
        project = get_object_or_404(Project, id=project_id)
        tasks = Task.objects.filter(projects=project).select_related('assignee', 'creator')
        
        # Create workbook
        wb = Workbook()
        
        # ===========================
        # SHEET 1: Summary
        # ===========================
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Styling
        header_fill = PatternFill(start_color="4F79FF", end_color="4F79FF", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=16, color="374151")
        
        # Title
        ws_summary['A1'] = f"Project Report: {project.project_name}"
        ws_summary['A1'].font = title_font
        ws_summary.merge_cells('A1:D1')
        
        # Date
        ws_summary['A2'] = f"Generated: {timezone.now().strftime('%B %d, %Y %H:%M')}"
        ws_summary['A2'].font = Font(size=10, color="6B7280")
        ws_summary.merge_cells('A2:D2')
        
        # Summary statistics
        ws_summary['A4'] = "Summary Statistics"
        ws_summary['A4'].font = Font(bold=True, size=14)
        
        total_tasks = tasks.count()
        total_done = tasks.filter(status="Done").count()
        overdue_count = tasks.filter(due_date__lt=timezone.now()).exclude(status="Done").count()
        completion_rate = round((total_done / total_tasks * 100) if total_tasks > 0 else 0, 1)
        
        stats_data = [
            ["Metric", "Value"],
            ["Total Tasks", total_tasks],
            ["Completed Tasks", total_done],
            ["Overdue Tasks", overdue_count],
            ["Completion Rate", f"{completion_rate}%"],
        ]
        
        for row_idx, row_data in enumerate(stats_data, start=6):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 6:  # Header row
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Adjust column widths
        ws_summary.column_dimensions['A'].width = 20
        ws_summary.column_dimensions['B'].width = 15
        
        # ===========================
        # SHEET 2: All Tasks
        # ===========================
        ws_tasks = wb.create_sheet(title="All Tasks")
        
        # Headers
        headers = ["Task Name", "Status", "Priority", "Assignee", "Creator", "Due Date", "Created At"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_tasks.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data
        for row_idx, task in enumerate(tasks, start=2):
            ws_tasks.cell(row=row_idx, column=1, value=task.task_name)
            ws_tasks.cell(row=row_idx, column=2, value=task.status)
            ws_tasks.cell(row=row_idx, column=3, value=task.priority)
            
            # Assignee name
            if task.assignee:
                assignee_name = f"{task.assignee.first_name} {task.assignee.last_name}" if task.assignee.first_name else task.assignee.email
            else:
                assignee_name = "Unassigned"
            ws_tasks.cell(row=row_idx, column=4, value=assignee_name)
            
            # Creator name
            if task.creator:
                creator_name = f"{task.creator.first_name} {task.creator.last_name}" if task.creator.first_name else task.creator.email
            else:
                creator_name = "N/A"
            ws_tasks.cell(row=row_idx, column=5, value=creator_name)
            
            ws_tasks.cell(row=row_idx, column=6, value=task.due_date.strftime('%Y-%m-%d') if task.due_date else "No deadline")
            ws_tasks.cell(row=row_idx, column=7, value=task.created_at.strftime('%Y-%m-%d %H:%M'))
            
            # Highlight overdue tasks
            if task.due_date and task.due_date < timezone.now() and task.status != "Done":
                for col in range(1, 8):
                    ws_tasks.cell(row=row_idx, column=col).fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        
        # Adjust column widths
        ws_tasks.column_dimensions['A'].width = 30
        ws_tasks.column_dimensions['B'].width = 12
        ws_tasks.column_dimensions['C'].width = 12
        ws_tasks.column_dimensions['D'].width = 20
        ws_tasks.column_dimensions['E'].width = 20
        ws_tasks.column_dimensions['F'].width = 12
        ws_tasks.column_dimensions['G'].width = 18
        
        # ===========================
        # SHEET 3: Overdue Tasks
        # ===========================
        overdue_tasks = tasks.filter(due_date__lt=timezone.now()).exclude(status="Done").order_by("due_date")
        
        if overdue_tasks.exists():
            ws_overdue = wb.create_sheet(title="Overdue Tasks")
            
            # Headers
            for col_idx, header in enumerate(["Task Name", "Assignee", "Due Date", "Days Overdue"], start=1):
                cell = ws_overdue.cell(row=1, column=col_idx, value=header)
                cell.fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Data
            for row_idx, task in enumerate(overdue_tasks, start=2):
                days_overdue = (timezone.now().date() - task.due_date.date()).days if hasattr(task.due_date, 'date') else (timezone.now().date() - task.due_date).days
                ws_overdue.cell(row=row_idx, column=1, value=task.task_name)
                
                # Assignee name
                if task.assignee:
                    assignee_name = f"{task.assignee.first_name} {task.assignee.last_name}" if task.assignee.first_name else task.assignee.email
                else:
                    assignee_name = "Unassigned"
                ws_overdue.cell(row=row_idx, column=2, value=assignee_name)
                
                ws_overdue.cell(row=row_idx, column=3, value=task.due_date.strftime('%Y-%m-%d'))
                ws_overdue.cell(row=row_idx, column=4, value=days_overdue)
            
            # Adjust column widths
            ws_overdue.column_dimensions['A'].width = 30
            ws_overdue.column_dimensions['B'].width = 20
            ws_overdue.column_dimensions['C'].width = 12
            ws_overdue.column_dimensions['D'].width = 15
        
        # ===========================
        # SHEET 4: Tasks by Status
        # ===========================
        ws_status = wb.create_sheet(title="By Status")
        
        # Count tasks by status
        statuses = ["Todo", "In Progress", "Done"]
        ws_status['A1'] = "Status Distribution"
        ws_status['A1'].font = Font(bold=True, size=14)
        ws_status.merge_cells('A1:B1')
        
        for row_idx, status in enumerate(statuses, start=3):
            count = tasks.filter(status=status).count()
            ws_status.cell(row=row_idx, column=1, value=status)
            ws_status.cell(row=row_idx, column=2, value=count)
            ws_status.cell(row=row_idx, column=1).font = Font(bold=True)
        
        ws_status.column_dimensions['A'].width = 15
        ws_status.column_dimensions['B'].width = 10
        
        # ===========================
        # Prepare response
        # ===========================
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"Report_{project.project_name.replace(' ', '_')}_{timezone.now().strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response

class SearchView(LoginRequiredMixin, TemplateView):
    template_name = "main_app/search_results.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        query = self.request.GET.get('q', '').strip()
        context['query'] = query
        
        if query:
            # Search Tasks
            tasks = Task.objects.filter(
                task_name__icontains=query
            ) | Task.objects.filter(
                task_description__icontains=query
            )
            
            # Search Projects
            projects = Project.objects.filter(
                project_name__icontains=query
            ) | Project.objects.filter(
                project_description__icontains=query
            )
            
            # Search Users
            users = User.objects.filter(
                first_name__icontains=query
            ) | User.objects.filter(
                last_name__icontains=query
            ) | User.objects.filter(
                email__icontains=query
            )
            
            context['tasks'] = tasks.distinct()[:20]  # Limit to 20 results
            context['projects'] = projects.distinct()[:20]
            context['users'] = users.distinct()[:20]
            
            # Count total results
            context['total_results'] = (
                tasks.distinct().count() + 
                projects.distinct().count() + 
                users.distinct().count()
            )
        else:
            context['tasks'] = []
            context['projects'] = []
            context['users'] = []
            context['total_results'] = 0
        
        return context

class AddCommentView(LoginRequiredMixin, View):
    def post(self, request, task_id):
        task = get_object_or_404(Task, id=task_id)
        user = request.user
        
        # Check if user is collaborator, assignee, or creator
        is_collaborator = (
            user == task.assignee or 
            user == task.creator or 
            user in task.collaborators.all()
        )
        
        if not is_collaborator:
            return HttpResponse("Forbidden: You must be a collaborator to comment", status=403)
        
        form = CommentForm(request.POST)
        if form.is_valid():
            comment = form.save(commit=False)
            comment.task = task
            comment.author = user
            comment.save()
        
        return redirect('main_app:one_task', task_id=task.id)

class DeleteCommentView(LoginRequiredMixin, View):
    def post(self, request, comment_id):
        comment = get_object_or_404(Comment, id=comment_id)
        task_id = comment.task.id
        
        # Only author can delete their own comment
        if comment.author != request.user:
            return HttpResponse("Forbidden: You can only delete your own comments", status=403)
        
        comment.delete()
        return redirect('main_app:one_task', task_id=task_id)
